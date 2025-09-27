#!/usr/bin/env python3
import re
import json
import openpyxl # New dependency for .xlsx files
from pathlib import Path
from typing import List, Dict, Optional, Union, Any, Tuple
from enum import Enum
from dataclasses import dataclass, field
from datetime import datetime
import csv # Added for a simpler BOM option if openpyxl isn't available
import argparse
import sys

# --- CONFIGURATION / FILE PATHS ---
# Set the specific file paths provided by the user
D356_NETLIST_FILE = "/Users/alaineperdomo/Desktop/IgniteHackathon/Test Case Files/UNO-TH_Rev3e_netlist.d356"
BOM_FILE = "/Users/alaineperdomo/Desktop/IgniteHackathon/Test Case Files/UNO-TH_Rev3e_bom.xlsx"
# The IPC file path from the original code is kept for completeness, 
# but the D356 and BOM files are the focus of the new 'parse_your_files'
IPC_FILE = "app/parser/Assembly Testpoint Report for Car-PCB1.ipc"


# --- DATA MODELS ---
class ComponentType(str, Enum):
    RESISTOR = "resistor"
    CAPACITOR = "capacitor"
    INDUCTOR = "inductor"
    DIODE = "diode"
    TRANSISTOR = "transistor"
    IC = "ic"
    CONNECTOR = "connector"
    CRYSTAL = "crystal"
    SWITCH = "switch"
    LED = "led"
    FUSE = "fuse"
    TRANSFORMER = "transformer"
    VIA = "via"
    TEST_POINT = "test_point"
    OTHER = "other"

class NetClass(str, Enum):
    POWER = "power"
    GROUND = "ground"
    SIGNAL = "signal"
    CLOCK = "clock"
    DIFFERENTIAL = "differential"
    HIGH_SPEED = "high_speed"
    ANALOG = "analog"
    OTHER = "other"

@dataclass
class Component:
    """Represents a component, combining netlist and BOM data."""
    reference: str
    # Data from Netlist/Placement
    x_position: Optional[float] = None
    y_position: Optional[float] = None
    rotation: Optional[float] = None
    layer: Optional[str] = None
    component_type: ComponentType = ComponentType.OTHER
    
    # Data from BOM
    value: Optional[str] = None # e.g., 10k, 0.1uF
    part_number: Optional[str] = None
    manufacturer: Optional[str] = None
    description: Optional[str] = None
    package: Optional[str] = None
    footprint: Optional[str] = None

@dataclass
class BOMEntry:
    """Temporary storage for data parsed from the BOM."""
    reference: str
    value: Optional[str] = None
    package: Optional[str] = None
    footprint: Optional[str] = None
    part_number: Optional[str] = None
    manufacturer: Optional[str] = None
    description: Optional[str] = None

@dataclass
class NetConnection:
    net_name: str
    component_ref: str
    pin_number: str
    pin_name: Optional[str] = None
    access_type: Optional[str] = None
    drill_size: Optional[float] = None
    x_position: Optional[float] = None
    y_position: Optional[float] = None

@dataclass
class Net:
    net_name: str
    connections: List[NetConnection]
    net_class: Optional[NetClass] = None

@dataclass
class ParsedNetlist:
    file_path: str
    file_format: str
    parse_timestamp: datetime
    nets: List[Net]
    components: List[Component]
    metadata: Dict[str, Any]
    
    @property
    def total_nets(self) -> int:
        return len(self.nets)
    
    @property
    def total_nodes(self) -> int:
        return sum(len(net.connections) for net in self.nets)

# --- HELPER FUNCTIONS ---
def infer_component_type(reference: str) -> ComponentType:
    """Infer component type from reference designator"""
    if not reference: return ComponentType.OTHER
    ref_upper = reference.upper()
    type_map = {
        'R': ComponentType.RESISTOR, 'C': ComponentType.CAPACITOR, 'L': ComponentType.INDUCTOR,
        'D': ComponentType.DIODE, 'Q': ComponentType.TRANSISTOR, 'U': ComponentType.IC,
        'IC': ComponentType.IC, 'J': ComponentType.CONNECTOR, 'P': ComponentType.CONNECTOR,
        'X': ComponentType.CRYSTAL, 'Y': ComponentType.CRYSTAL, 'S': ComponentType.SWITCH,
        'SW': ComponentType.SWITCH, 'LED': ComponentType.LED, 'F': ComponentType.FUSE,
        'T': ComponentType.TRANSFORMER, 'TP': ComponentType.TEST_POINT, 'VIA': ComponentType.VIA,
        'PTH': ComponentType.TEST_POINT,
    }
    for prefix, comp_type in type_map.items():
        if ref_upper.startswith(prefix):
            return comp_type
    return ComponentType.OTHER

def infer_net_class(net_name: str) -> NetClass:
    """Infer net class from net name"""
    if not net_name: return NetClass.OTHER
    name_upper = net_name.upper()
    # Power nets
    if any(power in name_upper for power in ['VCC', 'VDD', 'VIN', '+5V', '+3V3', '+12V', 'VBAT', 'VREF']):
        return NetClass.POWER
    # Ground nets
    if any(gnd in name_upper for gnd in ['GND', 'VSS', 'GROUND', 'EARTH']):
        return NetClass.GROUND
    # Clock signals
    if any(clk in name_upper for clk in ['CLK', 'CLOCK', 'OSC', 'XTAL']):
        return NetClass.CLOCK
    # Differential pairs
    if any(diff in name_upper for diff in ['_P', '_N', '+', '-']) and ('USB' in name_upper or 'DIFF' in name_upper):
        return NetClass.DIFFERENTIAL
    # High speed signals
    if any(hs in name_upper for hs in ['USB', 'ETHERNET', 'HDMI', 'PCIE']):
        return NetClass.HIGH_SPEED
    # Analog signals
    if any(analog in name_upper for analog in ['ADC', 'DAC', 'ANALOG', 'AREF', 'VREF']):
        return NetClass.ANALOG
    return NetClass.SIGNAL

# --- BOM PARSER ---
class BOMParser:
    """Parser for Bill of Materials (BOM) files (assumes .xlsx with common headers)."""
    def __init__(self):
        self.format_name = "BOM"
        # Common headers we look for (case-insensitive)
        self.header_map = {
            'designator': ['ref', 'reference', 'designator', 'component'],
            'value': ['value', 'val'],
            'part_number': ['part number', 'part_number', 'mfg part number', 'mfg_part_number'],
            'manufacturer': ['manufacturer', 'mfg'],
            'description': ['description', 'desc'],
            'package': ['package', 'footprint', 'pcb footprint']
        }

    def _find_headers(self, sheet) -> Dict[str, int]:
        """Identifies column indices for required fields."""
        headers = {}
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for idx, cell in enumerate(row):
                header_text = str(cell.value).strip().lower() if cell.value else ''
                for field_name, aliases in self.header_map.items():
                    if header_text in aliases:
                        headers[field_name] = idx + 1 # openpyxl is 1-indexed
                        break
        return headers

    def parse_file(self, file_path: Union[str, Path]) -> List[BOMEntry]:
        """Parses the BOM file and returns a list of BOMEntry objects."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"BOM file not found: {file_path}")

        print(f"Loading BOM from {file_path}")
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except Exception as e:
            # Fallback for non-xlsx BOMs, or if openpyxl fails
            print(f"Warning: Failed to load .xlsx with openpyxl ({e}). Trying as CSV...")
            return self._parse_csv_fallback(file_path)
            
        header_cols = self._find_headers(sheet)
        bom_entries: List[BOMEntry] = []

        if 'designator' not in header_cols:
            raise ValueError("Could not find 'Designator' column in BOM file.")

        # Iterate over rows starting from the second row (assuming row 1 is headers)
        for row in sheet.iter_rows(min_row=2):
            try:
                # Get the raw designator string, e.g., "R1, R2, R3"
                designator_cell = row[header_cols['designator'] - 1].value
                if not designator_cell:
                    continue

                designators = re.split(r'[,;\s]+', str(designator_cell).upper())
                
                # Fetch common data fields once per row
                row_data = {
                    field: str(row[col - 1].value).strip() if row[col - 1].value else None
                    for field, col in header_cols.items() if field != 'designator'
                }

                # Create an entry for each component reference
                for ref in designators:
                    if not ref: continue
                    bom_entries.append(BOMEntry(
                        reference=ref,
                        value=row_data.get('value'),
                        package=row_data.get('package'),
                        part_number=row_data.get('part_number'),
                        manufacturer=row_data.get('manufacturer'),
                        description=row_data.get('description'),
                    ))

            except IndexError:
                # This can happen if the row is shorter than expected, just skip
                continue
            except Exception as e:
                print(f"Warning: Error parsing BOM row: {e}")
                continue

        return bom_entries

    def _parse_csv_fallback(self, file_path: Path) -> List[BOMEntry]:
        """Simple CSV parser fallback if openpyxl fails or the file is CSV."""
        if file_path.suffix.lower() not in ['.csv', '.txt']:
             # Assume user specified an .xlsx but openpyxl failed, try to read it as a text file
             pass 

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            # Use 'sniff' to automatically detect the delimiter (comma, tab, whitespace)
            try:
                dialect = csv.Sniffer().sniff(f.read(1024))
                f.seek(0)
                reader = csv.reader(f, dialect)
            except csv.Error:
                f.seek(0)
                reader = csv.reader(f, delimiter='\t') # Default to tab if sniff fails
            
            rows = list(reader)
            if not rows: return []
            
            header_row = [str(h).strip().lower() for h in rows[0]]
            header_cols = {}
            for field_name, aliases in self.header_map.items():
                for alias in aliases:
                    if alias in header_row:
                        header_cols[field_name] = header_row.index(alias)
                        break
            
            if 'designator' not in header_cols:
                return [] # Cannot parse without designator

            bom_entries: List[BOMEntry] = []
            
            for row in rows[1:]:
                if not row: continue
                try:
                    designator_str = row[header_cols['designator']].upper()
                    designators = re.split(r'[,;\s]+', designator_str)
                    
                    row_data = {
                        field: row[header_cols[field]].strip() if header_cols[field] < len(row) else None
                        for field in header_cols if field != 'designator'
                    }

                    for ref in designators:
                        if not ref: continue
                        bom_entries.append(BOMEntry(
                            reference=ref,
                            value=row_data.get('value'),
                            package=row_data.get('package'),
                            part_number=row_data.get('part_number'),
                            manufacturer=row_data.get('manufacturer'),
                            description=row_data.get('description'),
                        ))

                except Exception:
                    continue
            return bom_entries


# --- D356 PARSER (Unchanged from original code) ---
class D356Parser:
    """D356 test file format parser"""
    # ... [D356Parser implementation remains the same as in the original code, 
    #       as its netlist parsing logic is correct.] ...
    def __init__(self):
        self.format_name = "D356"
        self.NET_RECORD = "317"
        self.ACCESS_RECORD = "327"
        self.PARAMETER = "P"
        self.COMMENT = "C"

    def parse_file(self, file_path: Union[str, Path]) -> ParsedNetlist:
        """Parse D356 test file"""
        file_path = Path(file_path)
        if not file_path.exists(): raise FileNotFoundError(f"D356 file not found: {file_path}")
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
        nets, components, metadata = self._parse_d356_content(content)
        return ParsedNetlist(file_path=str(file_path), file_format="D356", parse_timestamp=datetime.now(), nets=nets, components=components, metadata=metadata)

    def _parse_d356_content(self, content: str):
        """Parse D356 file content"""
        lines = content.strip().split('\n')
        nets_dict = {}
        components = {}
        metadata = {}
        for line_num, line in enumerate(lines, 1):
            if not line.strip(): continue
            try:
                if line.startswith(self.NET_RECORD): self._parse_net_record(line, nets_dict, components)
                elif line.startswith(self.ACCESS_RECORD): self._parse_access_record(line, nets_dict, components)
                elif line.startswith(self.COMMENT): self._parse_comment(line, metadata)
                elif line.startswith(self.PARAMETER): self._parse_parameter(line, metadata)
            except Exception as e: print(f"Warning: Error parsing line {line_num}: {e}"); continue
        nets = []
        for net_name, connections in nets_dict.items():
            if connections: nets.append(Net(net_name=net_name, connections=connections, net_class=infer_net_class(net_name)))
        return nets, list(components.values()), metadata

    def _parse_net_record(self, line: str, nets_dict: Dict, components: Dict):
        """Parse D356 317 record"""
        try:
            data = line[3:].strip()
            parts = re.split(r'\s{2,}', data)
            if len(parts) < 2: return
            net_name = parts[0].strip()
            if not net_name: return
            if net_name not in nets_dict: nets_dict[net_name] = []
            second_field = parts[1].strip()
            if second_field == "VIA": self._parse_via_record(net_name, parts[2:], nets_dict, components)
            else: self._parse_component_net_record(net_name, parts[1:], nets_dict, components)
        except Exception as e: print(f"Warning: Failed to parse net record: {e}")

    def _parse_via_record(self, net_name: str, drill_parts: List[str], nets_dict: Dict, components: Dict):
        """Parse VIA-based record"""
        if not drill_parts: return
        drill_info = ' '.join(drill_parts)
        pos_match = re.search(r'X([+-]?\d+)Y([+-]?\d+)', drill_info)
        if pos_match:
            try:
                x_pos = float(pos_match.group(1)) / 10000.0
                y_pos = float(pos_match.group(2)) / 10000.0
                via_ref = f"VIA_{net_name}_{len(nets_dict[net_name]) + 1}"
                connection = NetConnection(net_name=net_name, component_ref=via_ref, pin_number="1", access_type="via", x_position=x_pos, y_position=y_pos)
                nets_dict[net_name].append(connection)
                if via_ref not in components: components[via_ref] = Component(reference=via_ref, component_type=ComponentType.VIA, x_position=x_pos, y_position=y_pos, description=f"Via for {net_name}")
            except (ValueError, TypeError): pass

    def _parse_component_net_record(self, net_name: str, comp_parts: List[str], nets_dict: Dict, components: Dict):
        """Parse component-based record"""
        if not comp_parts: return
        comp_pin_field = comp_parts[0].strip()
        comp_pin_match = re.match(r'([A-Z0-9]+)\s*-\s*(\w*)', comp_pin_field)
        if comp_pin_match:
            component_ref = comp_pin_match.group(1)
            pin_number = comp_pin_match.group(2) if comp_pin_match.group(2) else "1"
            connection = NetConnection(net_name=net_name, component_ref=component_ref, pin_number=pin_number, access_type="component_pin")
            nets_dict[net_name].append(connection)
            if component_ref not in components: components[component_ref] = Component(reference=component_ref, component_type=infer_component_type(component_ref))

    def _parse_access_record(self, line: str, nets_dict: Dict, components: Dict):
        """Parse D356 327 record"""
        try:
            data = line[3:].strip()
            parts = re.split(r'\s{2,}', data)
            if len(parts) < 2: return
            net_name = parts[0].strip()
            comp_pin_field = parts[1].strip()
            comp_pin_match = re.match(r'([A-Z0-9]+)\s*-\s*(\w*)', comp_pin_field)
            if comp_pin_match:
                component_ref = comp_pin_match.group(1)
                pin_number = comp_pin_match.group(2) if comp_pin_match.group(2) else "1"
                if net_name not in nets_dict: nets_dict[net_name] = []
                connection = NetConnection(net_name=net_name, component_ref=component_ref, pin_number=pin_number, access_type="access_point")
                nets_dict[net_name].append(connection)
                if component_ref not in components: components[component_ref] = Component(reference=component_ref, component_type=infer_component_type(component_ref))
        except Exception as e: print(f"Warning: Failed to parse access record: {e}")

    def _parse_comment(self, line: str, metadata: Dict):
        """Parse comment lines"""
        comment_text = line[1:].strip()
        if "Project Name" in comment_text: metadata['project_name'] = comment_text.split(':')[-1].strip()
        elif "Board Name" in comment_text: metadata['board_name'] = comment_text.split(':')[-1].strip()
        elif "Date" in comment_text: metadata['date'] = comment_text.split(':')[-1].strip()

    def _parse_parameter(self, line: str, metadata: Dict):
        """Parse parameter lines"""
        param_data = line[1:].strip()
        parts = param_data.split(None, 1)
        if len(parts) >= 2: metadata[f'param_{parts[0].lower()}'] = parts[1].strip()

# --- IPC PARSER (Simplified/Placeholder) ---
class IPCParser:
    """Placeholder for IPC file parser."""
    def __init__(self): self.format_name = "IPC"
    def parse_file(self, file_path: Union[str, Path]) -> ParsedNetlist:
        # Simplified for brevity, as the focus is on BOM and D356
        return ParsedNetlist(file_path=str(file_path), file_format="IPC", parse_timestamp=datetime.now(), nets=[], components=[], metadata={})

# --- DATA MERGING FUNCTION ---
def reconcile_data(netlist_data: ParsedNetlist, bom_entries: List[BOMEntry]) -> ParsedNetlist:
    """
    Merges component data from the BOM into the component list from the netlist.
    """
    print("\nReconciling Netlist and BOM data...")
    # Create a dict of BOM entries keyed by component reference for fast lookup
    bom_dict = {entry.reference: entry for entry in bom_entries}
    
    # Components unique to Netlist (Netlist-only, usually test points, vias, or missing BOM data)
    netlist_only_count = 0
    # Components unique to BOM (BOM-only, usually bulk components, or components not placed on PCB)
    bom_only_count = 0 
    # Components successfully merged (found in both)
    merged_count = 0

    final_components: List[Component] = []
    
    # 1. Merge data for components found in the Netlist
    for comp in netlist_data.components:
        if comp.reference in bom_dict:
            # Found in both, merge BOM data into Netlist component
            bom_entry = bom_dict.pop(comp.reference) # Pop to track unmerged BOM entries
            comp.value = comp.value or bom_entry.value
            comp.package = comp.package or bom_entry.package
            comp.footprint = comp.footprint or bom_entry.footprint
            comp.part_number = comp.part_number or bom_entry.part_number
            comp.manufacturer = comp.manufacturer or bom_entry.manufacturer
            comp.description = comp.description or bom_entry.description
            merged_count += 1
        else:
            # Only in Netlist (e.g., VIAs, Test Points, unpopulated components with net connections)
            netlist_only_count += 1
        final_components.append(comp)

    # 2. Add remaining components from BOM that were not in the Netlist
    for ref, bom_entry in bom_dict.items():
        # Only in BOM (e.g., bulk items, materials, or simply unplaced components)
        bom_comp = Component(
            reference=ref,
            value=bom_entry.value,
            package=bom_entry.package,
            footprint=bom_entry.footprint,
            part_number=bom_entry.part_number,
            manufacturer=bom_entry.manufacturer,
            description=bom_entry.description,
            component_type=infer_component_type(ref)
        )
        final_components.append(bom_comp)
        bom_only_count += 1

    print(f"  - Merged components (Netlist & BOM): {merged_count}")
    print(f"  - Netlist-only components: {netlist_only_count}")
    print(f"  - BOM-only components: {bom_only_count}")
    print(f"  - Total components in final dataset: {len(final_components)}")
    
    netlist_data.components = final_components
    return netlist_data

def parse_radio_data(d356_file, bom_file, ipc_file=None, verbose=True):
    """
    Parse Radio PCB data from D356 netlist and BOM files.
    
    Args:
        d356_file (str): Path to the D356 netlist file
        bom_file (str): Path to the BOM Excel file
        ipc_file (str, optional): Path to the IPC file
        verbose (bool): Whether to print progress messages (default: True)
    
    Returns:
        dict: Parsed and reconciled PCB data
    """
    if verbose:
        print("--- STARTING PCB DATA STRUCTURING ---")
    
    results = { 'd356_data': None, 'ipc_data': None, 'bom_entries': [], 'analysis': {}, 'errors': [] }

    # Validate input files exist
    if not Path(d356_file).exists():
        error_msg = f"D356 file not found: {d356_file}"
        results['errors'].append(error_msg)
        if verbose:
            print(f"❌ ERROR: {error_msg}")
        return results

    if not Path(bom_file).exists():
        error_msg = f"BOM file not found: {bom_file}"
        results['errors'].append(error_msg)
        if verbose:
            print(f"❌ ERROR: {error_msg}")
        return results

    # 1. Parse D356 Netlist file
    if verbose:
        print(f"\n📁 1. Parsing D356 Netlist: {Path(d356_file).name}")
        print("-" * 50)
    try:
        d356_parser = D356Parser()
        d356_data = d356_parser.parse_file(d356_file)
        results['d356_data'] = d356_data
        if verbose:
            print(f"✅ SUCCESS: Parsed D356 file. {len(d356_data.nets)} nets, {len(d356_data.components)} components.")
    except Exception as e:
        error_msg = f"Failed to parse D356 Netlist file: {e}"
        results['errors'].append(error_msg)
        if verbose:
            print(f"❌ ERROR: {error_msg}")

    # 2. Parse BOM file
    if verbose:
        print(f"\n📄 2. Parsing BOM file: {Path(bom_file).name}")
        print("-" * 50)
    try:
        bom_parser = BOMParser()
        bom_entries = bom_parser.parse_file(bom_file)
        results['bom_entries'] = bom_entries
        if verbose:
            print(f"✅ SUCCESS: Parsed BOM file. {len(bom_entries)} component references found.")
    except Exception as e:
        error_msg = f"Failed to parse BOM file: {e}. Check if 'openpyxl' is installed (pip install openpyxl)."
        results['errors'].append(error_msg)
        if verbose:
            print(f"❌ ERROR: {error_msg}")

    # 3. Reconcile Data
    if results['d356_data'] and results['bom_entries']:
        if verbose:
            print(f"\n🔄 3. Reconciling Data")
            print("-" * 50)
        reconciled_data = reconcile_data(results['d356_data'], results['bom_entries'])
        results['d356_data'] = reconciled_data
        if verbose:
            print(f"✅ SUCCESS: Data reconciliation complete.")

    # 4. Parse IPC file (Optional)
    if ipc_file and Path(ipc_file).exists():
        if verbose:
            print(f"\n🔬 4. Parsing IPC file: {Path(ipc_file).name}")
            print("-" * 50)
        try:
            ipc_parser = IPCParser()
            ipc_data = ipc_parser.parse_file(ipc_file)
            results['ipc_data'] = ipc_data
            if verbose:
                print(f"✅ SUCCESS: Parsed IPC file.")
        except Exception as e:
            error_msg = f"Failed to parse IPC file: {e}"
            results['errors'].append(error_msg)
            if verbose:
                print(f"❌ ERROR: {error_msg}")
    elif ipc_file and verbose:
        print(f"ℹ️ IPC file not found at {ipc_file}. Skipping.")

    # 5. Analysis and Export (Using D356/Reconciled data as primary)
    if verbose:
        print(f"\n📈 5. Final Analysis and Export")
        print("-" * 50)
    
    # Run analysis logic on the reconciled D356 data
    if results['d356_data']:
        final_data = results['d356_data']
        if verbose:
            print(f"Final Structured Data Summary (D356 + BOM):")
            print(f" - Total Nets: {final_data.total_nets}")
            print(f" - Total Components: {len(final_data.components)}")
        
        component_types = {}
        for comp in final_data.components:
            comp_type = comp.component_type.value
            component_types[comp_type] = component_types.get(comp_type, 0) + 1
        
        net_classes = {}
        for net in final_data.nets:
            net_class = net.net_class.value if net.net_class else "unknown"
            net_classes[net_class] = net_classes.get(net_class, 0) + 1
        
        results['analysis']['component_types'] = component_types
        results['analysis']['net_classes'] = net_classes
        
        if verbose:
            print("Top Component Types:")
            for comp_type, count in sorted(component_types.items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f" • {comp_type}: {count}")
    
    if verbose:
        print("\n--- PCB DATA STRUCTURING COMPLETE ---")
    return results

def main():
    parser = argparse.ArgumentParser(
        description='Radio PCB Data Parser - Parse D356 netlist and BOM files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s netlist.d356 bom.xlsx
  %(prog)s netlist.d356 bom.xlsx --ipc testpoints.ipc
  %(prog)s /path/to/netlist.d356 /path/to/bom.xlsx --quiet
        """
    )
    
    parser.add_argument('d356_file', 
                       help='Path to the D356 netlist file')
    parser.add_argument('bom_file', 
                       help='Path to the BOM Excel file')
    parser.add_argument('--ipc', 
                       default=None,
                       help='Path to the IPC file (optional)')
    parser.add_argument('--quiet', '-q',
                       action='store_true',
                       help='Suppress verbose output')
    
    args = parser.parse_args()
    
    if not args.quiet:
        print("🚀 Radio PCB Data Parser")
        print("=" * 40)
        print(f"📁 D356 file: {args.d356_file}")
        print(f"📁 BOM file: {args.bom_file}")
        if args.ipc:
            print(f"📁 IPC file: {args.ipc}")
    
    try:
        result = parse_radio_data(
            args.d356_file, 
            args.bom_file, 
            ipc_file=args.ipc,
            verbose=not args.quiet
        )
        
        if result['errors']:
            if not args.quiet:
                print(f"❌ Completed with {len(result['errors'])} errors:")
                for error in result['errors']:
                    print(f"   - {error}")
            sys.exit(1)
        elif not args.quiet:
            print("✅ Successfully parsed radio PCB data!")
            
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()